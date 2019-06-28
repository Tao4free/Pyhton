#!/usr/local/bin/python3
# https://stackoverflow.com/questions/23058564/checking-a-character-is-fullwidth-or-halfwidth-in-python

import os
from openpyxl import load_workbook
import unicodedata as ud


'''
    For the given path, get the List of all files in the directory tree
'''
def getListOfFiles(dirName):
    # create a list of file and sub directories
    # names in the given directory
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)

    return allFiles

outdir = "output"
if not os.path.exists(outdir):
    os.makedirs(outdir)

indir = "input"
infiles = os.listdir(indir)

files = getListOfFiles(indir)
filenames = [f.split('/')[-1] for f in files]
#for fn in filenames:
#    print(fn)
any_ds = any("8-C-METI-07" in s for s in filenames)
#print(any_ds)

listfile = "list.xlsx"
wb = load_workbook(filename=listfile, read_only=True, data_only=True)
ws = wb["list"]
maxrow = ws.max_row

for n, row in enumerate(range(1,maxrow),1):
    iden = ws.cell(row, 3).value
    cond = ws.cell(row, 8).value
    fmt  = ws.cell(row,9).value
    if(cond is not None):
        cond_ds = '入手済' in cond
        if (n == 1):
            sfmt = "%-5s   %-11s   %-5s   %-10s"
        else:
            sfmt = "%-5s   %-13s   %-5s   %-10s"
        fmt_ds = True
        if (fmt is not None):
            fmt_ds = ud.east_asian_width(fmt[0]) == 'W'
        if (cond_ds is True and fmt_ds is False):
            any_ds = any(iden in s for s in filenames)
            if (not any_ds):
                fmt = fmt.replace('\n','&')
                out = sfmt % (n, iden, cond_ds, fmt)
                print(out)



