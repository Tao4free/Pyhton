#!/usr/bin/env python

# export data sheets from xlsx to csv

from openpyxl import load_workbook
import csv
from os import sys
import share

def get_all_sheets(excel_file):
    sheets = []
    workbook = load_workbook(excel_file,data_only=True)
    all_worksheets = workbook.get_sheet_names()
    for worksheet_name in all_worksheets:
        sheets.append(worksheet_name)
    return sheets

def csv_from_excel(excel_file, sheets):
    workbook = load_workbook(excel_file,data_only=True)
    for worksheet_name in sheets:
        #print("Export " + worksheet_name + " ...")

        try:
            worksheet = workbook.get_sheet_by_name(worksheet_name)
        except KeyError:
            #print("Could not find " + worksheet_name)
            sys.exit(1)

        dataitem_path = share.dataitem_path
        your_csv_file = open(''.join([dataitem_path, '/',worksheet_name,'.csv']), 'w',newline="\n")
        wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
        for row in worksheet.iter_rows():
            lrow = []
            for cell in row:
                lrow.append(cell.value)
            wr.writerow(lrow)
        #print(" ... done")
        your_csv_file.close()
