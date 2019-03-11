#!/usr/local/bin/python3

import os
from openpyxl import load_workbook

indir = "./"
infiles = os.listdir(indir)
inf = indir + '/' + "sample_added.xlsx"
inf_ref = indir + '/' + "reference.xlsx"

wb_ref = load_workbook(filename=inf_ref, read_only=True, data_only=True)
ws_ref = wb_ref["refer"]

wb = load_workbook(filename=inf)

start_no = 291
start_dis = 58.2

for row in range(1,ws_ref.max_row):
    nowrow = ws_ref.cell(row,1).value
    nextrow = ws_ref.cell(row+1,1).value
    if (nowrow is None and nextrow is None):
        break
    if ('k' in str(nowrow)):
        nowrow = str(nowrow).split('k')
        nowrow = nowrow[0]
        nowrow = round(float(nowrow),1)

        no = int( round((nowrow - start_dis),1) / 0.2 + start_no )
        noname = "No." + str(no)
        if (no < 291):
            continue
        if (no > 354):
            continue
        if (no%5 == 0):
            continue

        nownum = ws_ref.cell(row,2).value
        nextnum = ws_ref.cell(row+1,2).value

        #print(nowrow, ":  ", noname, ":  ", nownum, ":  ", nextnum)
        print(nowrow,no,noname)

        ws = wb[noname]

        for n in range(3, 3+nownum):
            dis = float(ws_ref.cell(row,n).value)
            dem = float(ws_ref.cell(row+1,n).value)
            if (n-2 == 1):
                dis = dis - 5
                dem = float(ws_ref.cell(row+1,n+1).value)
            if (n-2 == nownum):
                dis = dis + 5
                dem = float(ws_ref.cell(row+1,n-1).value)
            dis = round(dis,2)
            dem = round(dem,3)
            #print("num %2d %10.2f %10.3f" %  (n-2, dis, dem) )


            ws.cell(6 + n-2, 3).value = dis
            ws.cell(6 + n-2, 4).value = dem

        for m in range(7+nownum, 7+nownum+100):
             ws.cell(m, 3).value = None
             ws.cell(m, 4).value = None
                    


wb.save(inf)

print("Finish!", end='')

