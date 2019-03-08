#!/usr/local/bin/python3

import os
from openpyxl import load_workbook


def myfunc:
    for nrow in range(4,mxrow):
        #print(irow)
        for ncol in range(3,6):
            cellValue = sheet.cell(row=nrow, column=ncol).value
            if cellValue == None:
                return
            else:
                cellValue = round(cellValue, 4)
            if ncol == 3:
                #print("{0:10.2f}".format(cellValue), end='')
                cellValue = '%.2f' % cellValue
                str1 = str(cellValue)
            if ncol == 4:
                #print("{0:10.3f}".format(cellValue), end='') 
                cellValue = '%.3f' % cellValue
                str2 = str(cellValue)
            if ncol == 5:
                #print("{0:5d}".format(cellValue)) 
                cellValue = '%d' % cellValue
                str3 = str(cellValue)
        line     = '{:>10s} {:>10s} {:>5s}\n'.format(str1, str2, str3)
        f.write(line)


outdir = "output"
if not os.path.exists(outdir):
    os.makedirs(outdir)

indir = "input"
infiles = os.listdir(indir)

for inf in infiles:
    #print(inf)
    inf_path = indir + "/" + inf

    wb = load_workbook(filename=inf_path, read_only=True, data_only=True)
    
    for sheet in wb.worksheets:
        title = sheet.title
        mxrow = sheet.max_row
        print(inf, " | ", title)#,mxrow)
        
        of = outdir + '/' + title + '.txt'
        f = open(of, 'w')
    
        myfunc()
    
        f.close()

    wb._archive.close()

